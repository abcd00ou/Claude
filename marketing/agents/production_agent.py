"""
생산 에이전트 (Production Agent)
역할: BiCS5/6/8 NAND 생산 Capa 관리, 제품별 생산량 할당 시뮬레이션
산출물: Excel — production_simulation.xlsx
"""
import sys
sys.path.insert(0, str(__import__('pathlib').Path(__file__).parent.parent))

from pathlib import Path
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import SeriesLabel
import datetime

from config import PRODUCTS, NAND_COST_PER_GB, QUARTERS, EXCEL_DIR

# ── 생산 가정 데이터 ─────────────────────────────────────────
# BiCS 세대별 웨이퍼 Capa (K wafer/month) — WDC 기준 추정
NAND_WAFER_CAPA = {
    "BiCS5": {"capa_kwpm": 180, "die_per_wafer": 400, "gb_per_die": 32},  # 112L
    "BiCS6": {"capa_kwpm": 280, "die_per_wafer": 480, "gb_per_die": 64},  # 162L
    "BiCS8": {"capa_kwpm": 120, "die_per_wafer": 520, "gb_per_die": 96},  # 218L
}

# 분기별 Capa 성장률 가정
CAPA_GROWTH = {
    "BiCS5": -0.05,   # 단계적 축소 (구형 공정)
    "BiCS6":  0.08,   # 주력 확대
    "BiCS8":  0.20,   # 신공정 램프업
}

# 제품 카테고리별 NAND 사용 비중 (allocation %)
NAND_ALLOCATION = {
    "BiCS5": {"internal_ssd": 0.30, "external_ssd": 0.25, "microsd": 0.45},
    "BiCS6": {"internal_ssd": 0.40, "external_ssd": 0.40, "microsd": 0.20},
    "BiCS8": {"internal_ssd": 0.50, "external_ssd": 0.45, "microsd": 0.05},
}

# 제품별 NAND 투입량 (GB per unit)
NAND_INPUT_PER_UNIT = {sku: int(info["capacity_gb"] * 1.12)  # 12% overprovisioning
                       for sku, info in PRODUCTS.items()}

# ── 색상 팔레트 ─────────────────────────────────────────────
COLOR = {
    "header_dark":  "1B2A4A",   # 네이비
    "header_mid":   "D4001E",   # SanDisk 레드
    "bics5":        "4472C4",
    "bics6":        "ED7D31",
    "bics8":        "70AD47",
    "light_gray":   "F2F2F2",
    "white":        "FFFFFF",
    "yellow_flag":  "FFEB9C",
    "red_flag":     "FFC7CE",
    "green_ok":     "C6EFCE",
}

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color="000000", size=10):
    return Font(bold=bold, color=color, size=size, name="Calibri")

def _border_thin():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def _compute_quarterly_capa():
    """분기별 BiCS 세대별 생산 가능 GB (PetaByte 단위)"""
    result = {}
    base = {}
    for gen, spec in NAND_WAFER_CAPA.items():
        # 월간 GB = K wafer/month × 1000 × die/wafer × GB/die
        monthly_gb = spec["capa_kwpm"] * 1000 * spec["die_per_wafer"] * spec["gb_per_die"]
        base[gen] = monthly_gb * 3  # 분기 = 3개월

    for i, q in enumerate(QUARTERS):
        result[q] = {}
        for gen in NAND_WAFER_CAPA:
            growth = (1 + CAPA_GROWTH[gen]) ** i
            result[q][gen] = int(base[gen] * growth)
    return result


def build_excel():
    wb = openpyxl.Workbook()

    # ════════════════════════════════════════════════════════
    # 시트 1: Summary Dashboard
    # ════════════════════════════════════════════════════════
    ws_sum = wb.active
    ws_sum.title = "📊 Summary"

    ws_sum.column_dimensions["A"].width = 22
    for col in ["B","C","D","E","F","G","H","I"]:
        ws_sum.column_dimensions[col].width = 14
    ws_sum.row_dimensions[1].height = 40
    ws_sum.row_dimensions[2].height = 20

    # Title
    ws_sum.merge_cells("A1:I1")
    c = ws_sum["A1"]
    c.value = "SanDisk B2C Storage — 생산 Capa & 제품 할당 시뮬레이터"
    c.fill = _fill(COLOR["header_dark"])
    c.font = Font(bold=True, color="FFFFFF", size=14, name="Calibri")
    c.alignment = _center()

    ws_sum.merge_cells("A2:I2")
    c = ws_sum["A2"]
    c.value = f"생산 에이전트 | 기준일: {datetime.date.today()} | NAND: BiCS5 / BiCS6 / BiCS8"
    c.fill = _fill(COLOR["header_mid"])
    c.font = Font(bold=False, color="FFFFFF", size=9, name="Calibri")
    c.alignment = _center()

    # 분기별 총 Capa 헤더
    headers = ["BiCS 세대", "공정"] + QUARTERS
    ws_sum.append([])
    ws_sum.append(headers)
    row_h = ws_sum.max_row
    for col_idx, h in enumerate(headers, 1):
        c = ws_sum.cell(row=row_h, column=col_idx)
        c.fill = _fill(COLOR["header_dark"])
        c.font = _font(bold=True, color="FFFFFF", size=9)
        c.alignment = _center()
        c.border = _border_thin()

    quarterly_capa = _compute_quarterly_capa()
    gen_info = {"BiCS5": "112L TLC", "BiCS6": "162L TLC", "BiCS8": "218L TLC"}
    gen_colors = {"BiCS5": COLOR["bics5"], "BiCS6": COLOR["bics6"], "BiCS8": COLOR["bics8"]}

    for gen in ["BiCS5", "BiCS6", "BiCS8"]:
        row_data = [gen, gen_info[gen]]
        for q in QUARTERS:
            gb = quarterly_capa[q][gen]
            row_data.append(gb / 1e12)  # PB 단위
        ws_sum.append(row_data)
        r = ws_sum.max_row
        for col_idx in range(1, len(row_data)+1):
            c = ws_sum.cell(row=r, column=col_idx)
            c.border = _border_thin()
            c.alignment = _center()
            if col_idx <= 2:
                c.fill = _fill(gen_colors[gen])
                c.font = _font(bold=True, color="FFFFFF", size=9)
            else:
                c.fill = _fill(COLOR["light_gray"])
                c.font = _font(size=9)
                c.number_format = '#,##0.0 "PB"'

    # KPI 박스
    ws_sum.append([])
    ws_sum.append([])
    ws_sum.merge_cells(f"A{ws_sum.max_row}:I{ws_sum.max_row}")
    c = ws_sum.cell(row=ws_sum.max_row, column=1)
    c.value = "⚡ 주요 KPI (2025 연간 기준)"
    c.fill = _fill(COLOR["header_dark"])
    c.font = _font(bold=True, color="FFFFFF", size=11)
    c.alignment = _center()

    kpis = [
        ["총 BiCS6 생산량 (2025)", f"{sum(quarterly_capa[q]['BiCS6'] for q in QUARTERS[:4])/1e15:.1f} EB"],
        ["BiCS8 램프업 성장률", f"{CAPA_GROWTH['BiCS8']*100:.0f}% QoQ"],
        ["BiCS5 단계적 감축", f"{CAPA_GROWTH['BiCS5']*100:.0f}% QoQ"],
        ["평균 NAND 원가 (BiCS6)", f"${NAND_COST_PER_GB['BiCS6']:.3f}/GB"],
    ]
    for kpi in kpis:
        ws_sum.append(kpi)
        r = ws_sum.max_row
        ws_sum.cell(row=r, column=1).font = _font(bold=True, size=10)
        ws_sum.cell(row=r, column=2).font = _font(size=10, color="D4001E")
        ws_sum.cell(row=r, column=2).alignment = Alignment(horizontal="left")

    # ════════════════════════════════════════════════════════
    # 시트 2: NAND Capa Detail
    # ════════════════════════════════════════════════════════
    ws_capa = wb.create_sheet("🔬 NAND Capa")
    ws_capa.column_dimensions["A"].width = 20
    ws_capa.column_dimensions["B"].width = 14
    for col in ["C","D","E","F","G","H","I","J"]:
        ws_capa.column_dimensions[col].width = 15

    ws_capa.merge_cells("A1:J1")
    c = ws_capa["A1"]
    c.value = "NAND Capa 상세 — BiCS 세대별 분기별 생산 가능 출하량 (단위: TB)"
    c.fill = _fill(COLOR["header_dark"])
    c.font = Font(bold=True, color="FFFFFF", size=12, name="Calibri")
    c.alignment = _center()
    ws_capa.row_dimensions[1].height = 30

    headers2 = ["제품 카테고리", "NAND"] + QUARTERS
    ws_capa.append(headers2)
    r_h = ws_capa.max_row
    for ci, h in enumerate(headers2, 1):
        c = ws_capa.cell(row=r_h, column=ci)
        c.fill = _fill(COLOR["header_dark"])
        c.font = _font(bold=True, color="FFFFFF", size=9)
        c.alignment = _center()
        c.border = _border_thin()

    cats = ["internal_ssd", "external_ssd", "microsd"]
    cat_label = {"internal_ssd": "Internal SSD", "external_ssd": "External SSD", "microsd": "microSD"}

    for gen in ["BiCS5", "BiCS6", "BiCS8"]:
        for cat in cats:
            row_data = [cat_label[cat], gen]
            for q in QUARTERS:
                alloc_gb = quarterly_capa[q][gen] * NAND_ALLOCATION[gen][cat]
                row_data.append(alloc_gb / 1e9)  # TB 단위
            ws_capa.append(row_data)
            r = ws_capa.max_row
            for ci, val in enumerate(row_data, 1):
                c = ws_capa.cell(row=r, column=ci)
                c.border = _border_thin()
                c.alignment = _center()
                if ci == 1:
                    c.fill = _fill(COLOR["light_gray"])
                    c.font = _font(bold=True, size=9)
                elif ci == 2:
                    c.fill = _fill(gen_colors[gen])
                    c.font = _font(bold=True, color="FFFFFF", size=9)
                else:
                    c.font = _font(size=9)
                    c.number_format = '#,##0 "TB"'
        ws_capa.append([])  # 세대 간 구분

    # ════════════════════════════════════════════════════════
    # 시트 3: SKU별 생산 계획
    # ════════════════════════════════════════════════════════
    ws_sku = wb.create_sheet("📦 SKU 생산계획")
    ws_sku.column_dimensions["A"].width = 28
    ws_sku.column_dimensions["B"].width = 14
    ws_sku.column_dimensions["C"].width = 10
    ws_sku.column_dimensions["D"].width = 10
    for col in ["E","F","G","H","I","J","K","L"]:
        ws_sku.column_dimensions[col].width = 13

    ws_sku.merge_cells("A1:L1")
    c = ws_sku["A1"]
    c.value = "SKU별 분기 생산계획 (K units) — 수요예측 기반 초안"
    c.fill = _fill(COLOR["header_dark"])
    c.font = Font(bold=True, color="FFFFFF", size=12, name="Calibri")
    c.alignment = _center()
    ws_sku.row_dimensions[1].height = 30

    sku_headers = ["SKU", "카테고리", "NAND", "용량(GB)",
                   "원가($/unit)", "ASP($)", "GM%"] + QUARTERS[:4]
    ws_sku.append(sku_headers)
    r_h = ws_sku.max_row
    for ci, h in enumerate(sku_headers, 1):
        c = ws_sku.cell(row=r_h, column=ci)
        c.fill = _fill(COLOR["header_mid"])
        c.font = _font(bold=True, color="FFFFFF", size=9)
        c.alignment = _center()
        c.border = _border_thin()

    # SKU별 생산 계획 데이터 (K units/quarter — 수요 기반 추정)
    sku_plan = {
        # Internal SSD
        "WD_BLACK_SN850X_1TB":  [950, 1000, 1050, 1100],
        "WD_BLACK_SN850X_2TB":  [700, 750,  800,  850],
        "WD_BLACK_SN850X_4TB":  [200, 230,  260,  290],
        "WD_BLACK_SN770_1TB":   [600, 620,  640,  660],
        "WD_BLACK_SN770_2TB":   [400, 420,  440,  460],
        "SD_E30_1TB":           [450, 470,  490,  510],
        "SD_E30_2TB":           [280, 300,  320,  340],
        # External SSD
        "SD_EXTREME_1TB":       [1200,1250, 1300, 1350],
        "SD_EXTREME_2TB":       [800, 850,  900,  950],
        "SD_EXTREME_4TB":       [250, 280,  310,  340],
        "SD_EXTREME_PRO_1TB":   [500, 540,  580,  620],
        "SD_EXTREME_PRO_2TB":   [350, 380,  410,  440],
        "WD_MY_PASSPORT_1TB":   [900, 920,  940,  960],
        "WD_MY_PASSPORT_2TB":   [550, 570,  590,  610],
        # MicroSD
        "SD_EXTREME_MICRO_128G": [3000,3100,3200,3300],
        "SD_EXTREME_MICRO_256G": [4500,4700,4900,5100],
        "SD_EXTREME_MICRO_512G": [2800,2950,3100,3250],
        "SD_EXTREME_MICRO_1TB":  [1200,1350,1500,1650],
        "SD_ULTRA_MICRO_128G":   [5000,5200,5400,5600],
        "SD_ULTRA_MICRO_256G":   [4000,4200,4400,4600],
        "SD_ULTRA_MICRO_512G":   [2200,2350,2500,2650],
        "SD_PRO_PLUS_MICRO_256G":[1500,1600,1700,1800],
        "SD_PRO_PLUS_MICRO_512G":[1000,1100,1200,1300],
    }

    # ASP 추정 (USD)
    asp_map = {
        "WD_BLACK_SN850X_1TB":80,  "WD_BLACK_SN850X_2TB":145,  "WD_BLACK_SN850X_4TB":275,
        "WD_BLACK_SN770_1TB":65,   "WD_BLACK_SN770_2TB":115,
        "SD_E30_1TB":60,           "SD_E30_2TB":105,
        "SD_EXTREME_1TB":90,       "SD_EXTREME_2TB":160,       "SD_EXTREME_4TB":310,
        "SD_EXTREME_PRO_1TB":120,  "SD_EXTREME_PRO_2TB":210,
        "WD_MY_PASSPORT_1TB":75,   "WD_MY_PASSPORT_2TB":130,
        "SD_EXTREME_MICRO_128G":18,"SD_EXTREME_MICRO_256G":28, "SD_EXTREME_MICRO_512G":50,
        "SD_EXTREME_MICRO_1TB":95,
        "SD_ULTRA_MICRO_128G":12,  "SD_ULTRA_MICRO_256G":20,  "SD_ULTRA_MICRO_512G":38,
        "SD_PRO_PLUS_MICRO_256G":35,"SD_PRO_PLUS_MICRO_512G":65,
    }

    cat_order = ["internal_ssd", "external_ssd", "microsd"]
    cat_label_s = {"internal_ssd": "Internal SSD", "external_ssd": "External SSD", "microsd": "microSD"}
    last_cat = None

    for sku, info in sorted(PRODUCTS.items(), key=lambda x: cat_order.index(x[1]["cat"])):
        if info["cat"] != last_cat:
            # 카테고리 구분선
            ws_sku.append([f"▶ {cat_label_s[info['cat']]}"])
            r = ws_sku.max_row
            ws_sku.merge_cells(f"A{r}:L{r}")
            c = ws_sku.cell(row=r, column=1)
            c.fill = _fill(COLOR["header_dark"])
            c.font = _font(bold=True, color="FFFFFF", size=10)
            c.alignment = Alignment(horizontal="left", vertical="center")
            last_cat = info["cat"]

        nand_cost_unit = NAND_COST_PER_GB[info["nand"]] * info["capacity_gb"] * 1.12 + 8  # PCB/housing
        asp = asp_map.get(sku, 50)
        gm = (asp - nand_cost_unit) / asp * 100

        plan = sku_plan.get(sku, [0,0,0,0])
        row_data = [
            sku.replace("_", " "),
            cat_label_s[info["cat"]],
            info["nand"],
            info["capacity_gb"],
            round(nand_cost_unit, 2),
            asp,
            round(gm, 1),
        ] + plan

        ws_sku.append(row_data)
        r = ws_sku.max_row
        for ci, val in enumerate(row_data, 1):
            c = ws_sku.cell(row=r, column=ci)
            c.border = _border_thin()
            c.alignment = _center()
            c.font = _font(size=9)
            if ci == 7:  # GM%
                if gm >= 40:
                    c.fill = _fill(COLOR["green_ok"])
                elif gm >= 25:
                    c.fill = _fill(COLOR["yellow_flag"])
                else:
                    c.fill = _fill(COLOR["red_flag"])
                c.number_format = '0.0"%"'
            elif ci in [5,6]:
                c.number_format = '"$"#,##0.00'
            elif ci >= 8:
                c.number_format = '#,##0 "K"'
                c.fill = _fill(COLOR["light_gray"])

    # ════════════════════════════════════════════════════════
    # 시트 4: 원가 분석
    # ════════════════════════════════════════════════════════
    ws_cost = wb.create_sheet("💰 원가 분석")
    ws_cost.column_dimensions["A"].width = 25
    ws_cost.column_dimensions["B"].width = 12
    ws_cost.column_dimensions["C"].width = 12
    ws_cost.column_dimensions["D"].width = 12
    ws_cost.column_dimensions["E"].width = 15

    ws_cost.merge_cells("A1:E1")
    c = ws_cost["A1"]
    c.value = "NAND 원가 구조 분석 — BiCS 세대별 비교"
    c.fill = _fill(COLOR["header_dark"])
    c.font = Font(bold=True, color="FFFFFF", size=12, name="Calibri")
    c.alignment = _center()
    ws_cost.row_dimensions[1].height = 30

    cost_headers = ["항목", "BiCS5 (112L)", "BiCS6 (162L)", "BiCS8 (218L)", "비고"]
    ws_cost.append(cost_headers)
    r_h = ws_cost.max_row
    for ci, h in enumerate(cost_headers, 1):
        c = ws_cost.cell(row=r_h, column=ci)
        c.fill = _fill(COLOR["header_mid"])
        c.font = _font(bold=True, color="FFFFFF", size=9)
        c.alignment = _center()
        c.border = _border_thin()

    cost_data = [
        ["NAND $/GB", 0.055, 0.042, 0.038, "2025Q1 기준 추정"],
        ["1TB NAND 원가 ($)", 58.2, 44.4, 40.2, "12% OP 포함"],
        ["2TB NAND 원가 ($)", 116.4, 88.7, 80.3, ""],
        ["PCB/Controller/Housing ($)", 8.0, 8.0, 9.0, "BiCS8: 신규 ctrl"],
        ["Total BOM 1TB ($)", 66.2, 52.4, 49.2, ""],
        ["Total BOM 2TB ($)", 124.4, 96.7, 89.3, ""],
        ["외장 SSD 1TB ASP ($)", 90, 90, 120, "ExtPro = BiCS8"],
        ["외장 SSD 1TB GM% (%)", 26.4, 41.8, 59.0, "ASP 대비"],
        ["내장 SSD 1TB ASP ($)", 65, 80, "-", ""],
        ["내장 SSD 1TB GM% (%)", 1.8, 34.5, "-", "BiCS5 저마진"],
    ]
    for row in cost_data:
        ws_cost.append(row)
        r = ws_cost.max_row
        for ci, val in enumerate(row, 1):
            c = ws_cost.cell(row=r, column=ci)
            c.border = _border_thin()
            c.alignment = _center()
            c.font = _font(size=9)
            if ci == 1:
                c.fill = _fill(COLOR["light_gray"])
                c.font = _font(bold=True, size=9)

    # 차트 (BiCS 원가 비교)
    chart = BarChart()
    chart.type = "col"
    chart.title = "BiCS 세대별 1TB NAND 원가 ($)"
    chart.y_axis.title = "USD"
    chart.x_axis.title = "BiCS 세대"
    chart.style = 10
    chart.width = 15
    chart.height = 10

    data_ref = Reference(ws_cost, min_col=2, max_col=4, min_row=3, max_row=3)
    cats_ref = Reference(ws_cost, min_col=1, max_col=1, min_row=1, max_row=1)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    ws_cost.add_chart(chart, "A15")

    # 저장
    out_path = EXCEL_DIR / "production_simulation.xlsx"
    wb.save(out_path)
    print(f"[생산 에이전트] ✅ 저장: {out_path}")
    return out_path


if __name__ == "__main__":
    build_excel()
