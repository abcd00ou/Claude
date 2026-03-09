"""
공급 에이전트 (Supply Agent)
역할: 채널별 재고 관리, 공급 계획, SCM 시뮬레이션
산출물: Excel — supply_plan.xlsx
"""
import sys
sys.path.insert(0, str(__import__('pathlib').Path(__file__).parent.parent))

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, BarChart, Reference
import datetime

from config import PRODUCTS, QUARTERS, CHANNELS, REGIONS, EXCEL_DIR

# ── 공급 가정 데이터 ─────────────────────────────────────────
# 채널별 공급 비중 (카테고리별)
CHANNEL_MIX = {
    "internal_ssd": {"E-commerce": 0.55, "Retail": 0.30, "B2B/Corporate": 0.15},
    "external_ssd": {"E-commerce": 0.50, "Retail": 0.40, "B2B/Corporate": 0.10},
    "microsd":      {"E-commerce": 0.40, "Retail": 0.50, "B2B/Corporate": 0.10},
}

# 지역별 수요 비중
REGION_MIX = {
    "internal_ssd": {"NA":0.40, "EMEA":0.25, "APAC":0.20, "Japan":0.08, "China":0.07},
    "external_ssd": {"NA":0.35, "EMEA":0.28, "APAC":0.22, "Japan":0.08, "China":0.07},
    "microsd":      {"NA":0.25, "EMEA":0.30, "APAC":0.25, "Japan":0.10, "China":0.10},
}

# Lead time (weeks)
LEAD_TIME = {"E-commerce": 2, "Retail": 5, "B2B/Corporate": 8}

# Target Inventory Days (DOI)
TARGET_DOI = {"E-commerce": 21, "Retail": 45, "B2B/Corporate": 30}

# 분기별 총 출하 계획 (K units — production_agent에서 연동)
TOTAL_SHIPMENT_PLAN = {
    "internal_ssd": {"2025Q1":3580, "2025Q2":3790, "2025Q3":4000, "2025Q4":4210,
                     "2026Q1":4400, "2026Q2":4600, "2026Q3":4800, "2026Q4":5000},
    "external_ssd": {"2025Q1":4550, "2025Q2":4790, "2025Q3":5040, "2025Q4":5270,
                     "2026Q1":5500, "2026Q2":5750, "2026Q3":6010, "2026Q4":6280},
    "microsd":      {"2025Q1":25200,"2025Q2":26800,"2025Q3":28400,"2025Q4":30000,
                     "2026Q1":31200,"2026Q2":32500,"2026Q3":33800,"2026Q4":35100},
}

# 재고 위험 임계값
STOCKOUT_RISK_DOI = 15   # DOI < 15일 = 재고 부족 경고
OVERSTOCK_DOI     = 75   # DOI > 75일 = 과잉재고 경고

COLOR = {
    "header_dark": "1B2A4A",
    "header_mid":  "D4001E",
    "light_gray":  "F2F2F2",
    "green_ok":    "C6EFCE",
    "yellow_warn": "FFEB9C",
    "red_risk":    "FFC7CE",
    "white":       "FFFFFF",
    "na":          "4472C4",
    "emea":        "ED7D31",
    "apac":        "70AD47",
    "japan":       "9E0A0A",
    "china":       "FFC000",
}

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color="000000", size=10):
    return Font(bold=bold, color=color, size=size, name="Calibri")

def _border_thin():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def build_excel():
    wb = openpyxl.Workbook()

    # ════════════════════════════════════════════════════════
    # 시트 1: 채널별 공급 계획
    # ════════════════════════════════════════════════════════
    ws_ch = wb.active
    ws_ch.title = "📦 채널별 공급계획"

    ws_ch.column_dimensions["A"].width = 18
    ws_ch.column_dimensions["B"].width = 18
    for col in ["C","D","E","F","G","H","I","J"]:
        ws_ch.column_dimensions[col].width = 13

    ws_ch.merge_cells("A1:J1")
    c = ws_ch["A1"]
    c.value = "SanDisk B2C — 채널별 분기 공급 계획 (K units)"
    c.fill = _fill(COLOR["header_dark"])
    c.font = Font(bold=True, color="FFFFFF", size=13, name="Calibri")
    c.alignment = _center()
    ws_ch.row_dimensions[1].height = 32

    headers = ["카테고리", "채널"] + QUARTERS
    ws_ch.append(headers)
    rh = ws_ch.max_row
    for ci, h in enumerate(headers, 1):
        c = ws_ch.cell(row=rh, column=ci)
        c.fill = _fill(COLOR["header_mid"])
        c.font = _font(bold=True, color="FFFFFF", size=9)
        c.alignment = _center()
        c.border = _border_thin()

    cat_label = {"internal_ssd":"Internal SSD","external_ssd":"External SSD","microsd":"microSD"}
    ch_colors = {"E-commerce":"4472C4","Retail":"ED7D31","B2B/Corporate":"70AD47"}

    for cat in ["internal_ssd","external_ssd","microsd"]:
        # 카테고리 구분
        ws_ch.append([f"▶ {cat_label[cat]}"])
        r = ws_ch.max_row
        ws_ch.merge_cells(f"A{r}:J{r}")
        c = ws_ch.cell(row=r, column=1)
        c.fill = _fill(COLOR["header_dark"])
        c.font = _font(bold=True, color="FFFFFF", size=10)
        c.alignment = Alignment(horizontal="left", vertical="center")

        for ch in CHANNELS:
            row_data = [cat_label[cat], ch]
            for q in QUARTERS:
                total = TOTAL_SHIPMENT_PLAN[cat][q]
                alloc = total * CHANNEL_MIX[cat][ch]
                row_data.append(round(alloc, 0))
            ws_ch.append(row_data)
            r = ws_ch.max_row
            for ci, val in enumerate(row_data, 1):
                c = ws_ch.cell(row=r, column=ci)
                c.border = _border_thin()
                c.alignment = _center()
                if ci == 2:
                    c.fill = _fill(ch_colors[ch])
                    c.font = _font(bold=True, color="FFFFFF", size=9)
                elif ci >= 3:
                    c.fill = _fill(COLOR["light_gray"])
                    c.font = _font(size=9)
                    c.number_format = '#,##0 "K"'
                else:
                    c.font = _font(size=9)

    # ════════════════════════════════════════════════════════
    # 시트 2: 재고 시뮬레이션 (DOI)
    # ════════════════════════════════════════════════════════
    ws_doi = wb.create_sheet("📊 재고 DOI 시뮬")

    ws_doi.column_dimensions["A"].width = 18
    ws_doi.column_dimensions["B"].width = 18
    for col in ["C","D","E","F","G","H","I","J"]:
        ws_doi.column_dimensions[col].width = 13

    ws_doi.merge_cells("A1:J1")
    c = ws_doi["A1"]
    c.value = "재고 Days on Hand (DOI) 시뮬레이션 — 채널별 위험도 평가"
    c.fill = _fill(COLOR["header_dark"])
    c.font = Font(bold=True, color="FFFFFF", size=12, name="Calibri")
    c.alignment = _center()
    ws_doi.row_dimensions[1].height = 30

    ws_doi.append([])
    legend_row = ws_doi.max_row + 1
    ws_doi.append(["범례:", "✅ 정상 (21~60일)", "⚠️ 위험 (<15일 또는 >75일)", "",
                   f"Stockout 기준: <{STOCKOUT_RISK_DOI}일", f"과잉재고 기준: >{OVERSTOCK_DOI}일"])
    ws_doi.cell(row=ws_doi.max_row, column=2).fill = _fill(COLOR["green_ok"])
    ws_doi.cell(row=ws_doi.max_row, column=3).fill = _fill(COLOR["red_risk"])

    doi_headers = ["카테고리", "채널"] + QUARTERS
    ws_doi.append(doi_headers)
    rh = ws_doi.max_row
    for ci, h in enumerate(doi_headers, 1):
        c = ws_doi.cell(row=rh, column=ci)
        c.fill = _fill(COLOR["header_mid"])
        c.font = _font(bold=True, color="FFFFFF", size=9)
        c.alignment = _center()
        c.border = _border_thin()

    # DOI 시뮬레이션 (분기별 계절성 반영)
    SEASONAL = {
        "2025Q1": 0.85, "2025Q2": 0.95, "2025Q3": 1.10, "2025Q4": 1.30,
        "2026Q1": 0.88, "2026Q2": 0.97, "2026Q3": 1.08, "2026Q4": 1.28,
    }

    for cat in ["internal_ssd","external_ssd","microsd"]:
        ws_doi.append([f"▶ {cat_label[cat]}"])
        r = ws_doi.max_row
        ws_doi.merge_cells(f"A{r}:J{r}")
        c = ws_doi.cell(row=r, column=1)
        c.fill = _fill(COLOR["header_dark"])
        c.font = _font(bold=True, color="FFFFFF", size=10)
        c.alignment = Alignment(horizontal="left", vertical="center")

        for ch in CHANNELS:
            target_doi = TARGET_DOI[ch]
            row_data = [cat_label[cat], ch]
            for q in QUARTERS:
                # 계절성으로 인해 Q4에 DOI 감소 (수요 급증)
                season = SEASONAL[q]
                simulated_doi = target_doi / season * (0.9 + 0.2 * (CHANNELS.index(ch) * 0.1))
                row_data.append(round(simulated_doi, 1))
            ws_doi.append(row_data)
            r = ws_doi.max_row
            for ci, val in enumerate(row_data, 1):
                c = ws_doi.cell(row=r, column=ci)
                c.border = _border_thin()
                c.alignment = _center()
                c.font = _font(size=9)
                if ci >= 3:
                    doi_val = val
                    if doi_val < STOCKOUT_RISK_DOI:
                        c.fill = _fill(COLOR["red_risk"])
                    elif doi_val > OVERSTOCK_DOI:
                        c.fill = _fill(COLOR["yellow_warn"])
                    else:
                        c.fill = _fill(COLOR["green_ok"])
                    c.number_format = '0.0 "일"'

    # ════════════════════════════════════════════════════════
    # 시트 3: 지역별 출하 계획
    # ════════════════════════════════════════════════════════
    ws_reg = wb.create_sheet("🌏 지역별 출하")

    ws_reg.column_dimensions["A"].width = 18
    ws_reg.column_dimensions["B"].width = 10
    for col in ["C","D","E","F","G","H","I","J"]:
        ws_reg.column_dimensions[col].width = 13

    ws_reg.merge_cells("A1:J1")
    c = ws_reg["A1"]
    c.value = "지역별 분기 출하 계획 (K units) — NA / EMEA / APAC / Japan / China"
    c.fill = _fill(COLOR["header_dark"])
    c.font = Font(bold=True, color="FFFFFF", size=12, name="Calibri")
    c.alignment = _center()
    ws_reg.row_dimensions[1].height = 30

    reg_headers = ["카테고리", "지역"] + QUARTERS
    ws_reg.append(reg_headers)
    rh = ws_reg.max_row
    for ci, h in enumerate(reg_headers, 1):
        c = ws_reg.cell(row=rh, column=ci)
        c.fill = _fill(COLOR["header_mid"])
        c.font = _font(bold=True, color="FFFFFF", size=9)
        c.alignment = _center()
        c.border = _border_thin()

    reg_colors = {r: COLOR.get(r.lower(), "888888") for r in REGIONS}

    for cat in ["internal_ssd","external_ssd","microsd"]:
        ws_reg.append([f"▶ {cat_label[cat]}"])
        r_sep = ws_reg.max_row
        ws_reg.merge_cells(f"A{r_sep}:J{r_sep}")
        c = ws_reg.cell(row=r_sep, column=1)
        c.fill = _fill(COLOR["header_dark"])
        c.font = _font(bold=True, color="FFFFFF", size=10)
        c.alignment = Alignment(horizontal="left", vertical="center")

        for reg in REGIONS:
            row_data = [cat_label[cat], reg]
            for q in QUARTERS:
                total = TOTAL_SHIPMENT_PLAN[cat][q]
                alloc = total * REGION_MIX[cat][reg]
                row_data.append(round(alloc, 0))
            ws_reg.append(row_data)
            r = ws_reg.max_row
            for ci, val in enumerate(row_data, 1):
                c = ws_reg.cell(row=r, column=ci)
                c.border = _border_thin()
                c.alignment = _center()
                c.font = _font(size=9)
                if ci == 2:
                    c.fill = _fill(reg_colors[reg])
                    c.font = _font(bold=True, color="FFFFFF", size=9)
                elif ci >= 3:
                    c.fill = _fill(COLOR["light_gray"])
                    c.number_format = '#,##0 "K"'

    # ════════════════════════════════════════════════════════
    # 시트 4: SCM 리스크 대시보드
    # ════════════════════════════════════════════════════════
    ws_risk = wb.create_sheet("⚠️ SCM 리스크")

    ws_risk.column_dimensions["A"].width = 28
    ws_risk.column_dimensions["B"].width = 15
    ws_risk.column_dimensions["C"].width = 15
    ws_risk.column_dimensions["D"].width = 15
    ws_risk.column_dimensions["E"].width = 30

    ws_risk.merge_cells("A1:E1")
    c = ws_risk["A1"]
    c.value = "SCM 리스크 레지스터 (2025년 기준)"
    c.fill = _fill(COLOR["header_dark"])
    c.font = Font(bold=True, color="FFFFFF", size=12, name="Calibri")
    c.alignment = _center()
    ws_risk.row_dimensions[1].height = 30

    risk_headers = ["리스크 항목", "발생 가능성", "영향도", "총 점수", "대응 방안"]
    ws_risk.append(risk_headers)
    rh = ws_risk.max_row
    for ci, h in enumerate(risk_headers, 1):
        c = ws_risk.cell(row=rh, column=ci)
        c.fill = _fill(COLOR["header_mid"])
        c.font = _font(bold=True, color="FFFFFF", size=9)
        c.alignment = _center()
        c.border = _border_thin()

    risks = [
        ["NAND 가격 급등 (공급 타이트)", "중(3)", "상(4)", 12, "BiCS6 Capa 선제 확대, 장기 계약 체결"],
        ["Q4 계절 수요 초과 (재고 부족)", "상(4)", "상(4)", 16, "Q3 선행 생산 확대, Safety stock +20%"],
        ["경쟁사 가격 인하 공세", "상(4)", "중(3)", 12, "가격 모니터링 주간, 프로모션 예산 유보"],
        ["물류 리드타임 증가 (항운 파업)", "하(2)", "상(4)", 8,  "항공 긴급 전환 옵션 유지"],
        ["Amazon 알고리즘 변경 노출도↓", "중(3)", "중(3)", 9,  "SEO 개선, Sponsored 예산 상향"],
        ["BiCS8 양품률 초기 저조", "중(3)", "중(3)", 9,  "BiCS6 병행 유지, BiCS8 ramping 보수적"],
        ["환율 변동 (JPY/EUR 약세)", "중(3)", "하(2)", 6,  "환헤지 계약 (FWD 3개월 단위)"],
        ["MicroSD 중국산 저가 공세", "상(4)", "중(3)", 12, "ProPlus 차별화, 브랜드 프리미엄 강화"],
    ]

    for risk in risks:
        ws_risk.append(risk)
        r = ws_risk.max_row
        score = risk[3]
        for ci, val in enumerate(risk, 1):
            c = ws_risk.cell(row=r, column=ci)
            c.border = _border_thin()
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.font = _font(size=9)
            if ci == 4:
                if score >= 12:
                    c.fill = _fill(COLOR["red_risk"])
                    c.font = _font(bold=True, size=9)
                elif score >= 8:
                    c.fill = _fill(COLOR["yellow_warn"])
                else:
                    c.fill = _fill(COLOR["green_ok"])
        ws_risk.row_dimensions[r].height = 25

    out_path = EXCEL_DIR / "supply_plan.xlsx"
    wb.save(out_path)
    print(f"[공급 에이전트] ✅ 저장: {out_path}")
    return out_path


if __name__ == "__main__":
    build_excel()
